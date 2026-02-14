"""
MemoryCharm Financial Model Generator  v2
Creates an Excel workbook with adjustable assumptions and linked projections.
Now with granular infrastructure: Azure Blob Cool, Cloudflare R2, Table Storage,
Functions invocations, content-size-by-type, and playback request volume.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styling ──────────────────────────────────────────────────────────────────

DARK_BLUE = "1B2A4A"
MED_BLUE = "2E4057"
LIGHT_BLUE = "D6E4F0"
ACCENT_GOLD = "C9A961"
WHITE = "FFFFFF"
MED_GRAY = "D9D9D9"
GREEN_FILL = "E2EFDA"
RED_FILL = "FCE4D6"
INPUT_FILL = "FFF2CC"

header_font = Font(name="Calibri", bold=True, size=12, color=WHITE)
header_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
input_fill_style = PatternFill(start_color=INPUT_FILL, end_color=INPUT_FILL, fill_type="solid")
normal_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", bold=True, size=11)
title_font = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
section_font = Font(name="Calibri", bold=True, size=13, color=DARK_BLUE)
note_font = Font(name="Calibri", italic=True, size=9, color="888888")
currency_fmt = '$#,##0.00'
currency_whole = '$#,##0'
currency_micro = '$#,##0.0000'
pct_fmt = '0.0%'
num_fmt = '#,##0'
num_1dp = '#,##0.0'
num_2dp = '#,##0.00'
thin_border = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)
green_fill = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type="solid")
red_fill = PatternFill(start_color=RED_FILL, end_color=RED_FILL, fill_type="solid")
blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")


def style_range(ws, row, col_start, col_end, font=None, fill=None, number_format=None, alignment=None, border=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if number_format:
            cell.number_format = number_format
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border


def sc(ws, row, col, value, font=None, fill=None, number_format=None, alignment=None, border=None):
    """Set cell with styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if number_format: cell.number_format = number_format
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    return cell


def inp(ws, row, col, value, number_format=None):
    """Editable assumption cell (yellow)."""
    return sc(ws, row, col, value, font=bold_font, fill=input_fill_style, number_format=number_format, border=thin_border)


def note(ws, row, col, text):
    """Italic gray note."""
    return sc(ws, row, col, text, font=note_font)


def formula(ws, row, col, f, font=None, number_format=None, fill=None):
    """Set a formula cell."""
    cell = ws.cell(row=row, column=col)
    cell.value = f
    cell.font = font or normal_font
    cell.number_format = number_format or num_fmt
    cell.border = thin_border
    if fill:
        cell.fill = fill
    return cell


def section_header(ws, row, col_start, col_end, headers):
    """Dark blue header row."""
    for i, h in enumerate(headers):
        sc(ws, row, col_start + i, h, font=header_font, fill=header_fill, alignment=Alignment(horizontal="center"))


# ── Workbook ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Assumptions"
ws.sheet_properties.tabColor = DARK_BLUE
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 48
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 40

sc(ws, 1, 2, "MemoryCharm Financial Model — Assumptions", font=title_font)
sc(ws, 2, 2, "Yellow cells are editable — change them to see impact on Projections & Summary sheets", font=Font(name="Calibri", italic=True, size=10, color="666666"))

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Charm Pricing  (rows 4-8)
# ────────────────────────────────────────────────────────────────────────────
R = 4
sc(ws, R, 2, "CHARM PRICING (Retail to Consumer)", font=section_font)
R += 1  # 5
section_header(ws, R, 2, 5, ["Tier", "Unit Price", "COGS / Unit", "Gross Margin"])
R += 1  # 6 — 10-Year
sc(ws, R, 2, "10-Year Charm", font=normal_font, border=thin_border)
inp(ws, R, 3, 29.99, currency_fmt); inp(ws, R, 4, 8.50, currency_fmt)
formula(ws, R, 5, "=C6-D6", bold_font, currency_fmt)
R += 1  # 7 — 15-Year
sc(ws, R, 2, "15-Year Charm", font=normal_font, border=thin_border)
inp(ws, R, 3, 44.99, currency_fmt); inp(ws, R, 4, 8.50, currency_fmt)
formula(ws, R, 7, "=C7-D7", bold_font, currency_fmt)
R += 1  # 8 — Perpetual
sc(ws, R, 2, "Retail (Perpetual) Charm", font=normal_font, border=thin_border)
inp(ws, R, 3, 69.99, currency_fmt); inp(ws, R, 4, 9.00, currency_fmt)
formula(ws, R, 5, "=C8-D8", bold_font, currency_fmt)

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Upsell Pricing  (rows 10-14)
# ────────────────────────────────────────────────────────────────────────────
R = 10
sc(ws, R, 2, "UPSELL / ADD-ON PRICING", font=section_font)
R += 1  # 11
section_header(ws, R, 2, 5, ["Product", "Price", "Cost", "Margin"])
R += 1  # 12
sc(ws, R, 2, "Extend Memory (+5 years)", font=normal_font, border=thin_border)
inp(ws, R, 3, 14.99, currency_fmt); inp(ws, R, 4, 1.00, currency_fmt)
formula(ws, R, 5, "=C12-D12", bold_font, currency_fmt)
R += 1  # 13
sc(ws, R, 2, "Upgrade Tier (avg revenue per upgrade)", font=normal_font, border=thin_border)
inp(ws, R, 3, 19.99, currency_fmt); inp(ws, R, 4, 0.50, currency_fmt)
formula(ws, R, 5, "=C13-D13", bold_font, currency_fmt)
R += 1  # 14
sc(ws, R, 2, "Gift Wrap", font=normal_font, border=thin_border)
inp(ws, R, 3, 4.99, currency_fmt); inp(ws, R, 4, 1.50, currency_fmt)
formula(ws, R, 5, "=C14-D14", bold_font, currency_fmt)

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Sales Volume  (rows 16-20)
# ────────────────────────────────────────────────────────────────────────────
R = 16
sc(ws, R, 2, "MONTHLY UNIT SALES (Starting Month 1)", font=section_font)
R += 1  # 17
section_header(ws, R, 2, 4, ["Item", "Units / Month", "MoM Growth %"])
R += 1  # 18
sc(ws, R, 2, "10-Year Charms", font=normal_font, border=thin_border)
inp(ws, R, 3, 100, num_fmt); inp(ws, R, 4, 0.08, pct_fmt)
R += 1  # 19
sc(ws, R, 2, "15-Year Charms", font=normal_font, border=thin_border)
inp(ws, R, 3, 40, num_fmt); inp(ws, R, 4, 0.10, pct_fmt)
R += 1  # 20
sc(ws, R, 2, "Retail (Perpetual) Charms", font=normal_font, border=thin_border)
inp(ws, R, 3, 15, num_fmt); inp(ws, R, 4, 0.12, pct_fmt)

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Upsell Attach Rates  (rows 22-25)
# ────────────────────────────────────────────────────────────────────────────
R = 22
sc(ws, R, 2, "UPSELL ATTACH RATES", font=section_font)
R += 1  # 23
sc(ws, R, 2, "Extend Memory (monthly % of cumulative base)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.02, pct_fmt); note(ws, R, 4, "Applied to cumulative charms sold")
R += 1  # 24
sc(ws, R, 2, "Upgrade Tier (monthly % of cumulative base)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.01, pct_fmt)
R += 1  # 25
sc(ws, R, 2, "Gift Wrap (% of new monthly charm sales)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.15, pct_fmt)

# ════════════════════════════════════════════════════════════════════════════
# SECTION: CONTENT SIZE ASSUMPTIONS  (rows 27-35)
# ════════════════════════════════════════════════════════════════════════════
R = 27
sc(ws, R, 2, "CONTENT SIZE ASSUMPTIONS", font=section_font)
note(ws, R, 6, "API enforces 150 MB max per charm")

R += 1  # 28
section_header(ws, R, 2, 5, ["Media Type", "Avg Size (MB)", "% of Charms", "Weighted MB"])
note(ws, R, 6, "Weighted MB = Size x Mix %")

R += 1  # 29 — Video
sc(ws, R, 2, "Video charms (single mp4/webm/mov)", font=normal_font, border=thin_border)
inp(ws, R, 3, 55, num_fmt)
inp(ws, R, 4, 0.55, pct_fmt)
formula(ws, R, 5, "=C29*D29", normal_font, num_2dp)
note(ws, R, 6, "iPhone 1080p: HEVC ~20MB, H.264 ~65MB, 4K ~67MB for 30s")

R += 1  # 30 — Image gallery
sc(ws, R, 2, "Image gallery charms (1-20 images)", font=normal_font, border=thin_border)
inp(ws, R, 3, 15, num_fmt)
inp(ws, R, 4, 0.35, pct_fmt)
formula(ws, R, 5, "=C30*D30", normal_font, num_2dp)
note(ws, R, 6, "~4-6 iPhone photos avg at 2.5-4 MB each (HEIF/JPEG)")

R += 1  # 31 — Audio
sc(ws, R, 2, "Audio charms (single mp3/wav/ogg/aac)", font=normal_font, border=thin_border)
inp(ws, R, 3, 6, num_fmt)
inp(ws, R, 4, 0.10, pct_fmt)
formula(ws, R, 5, "=C31*D31", normal_font, num_2dp)
note(ws, R, 6, "Voice message or song clip, 1-3 min")

R += 1  # 32 — Mix validation
sc(ws, R, 2, "Mix total (should = 100%)", font=bold_font, border=thin_border)
formula(ws, R, 4, "=D29+D30+D31", bold_font, pct_fmt)

R += 1  # 33 — Weighted average
sc(ws, R, 2, "WEIGHTED AVG CONTENT SIZE (MB per charm)", font=bold_font, border=thin_border)
formula(ws, R, 3, "=E29+E30+E31", bold_font, num_2dp)
note(ws, R, 6, "Used for all storage cost calculations below")
sc(ws, R, 5, None, font=bold_font, border=thin_border)

R += 1  # 34 — Claim rate
sc(ws, R, 2, "Charm claim rate (% sold that get activated)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.85, pct_fmt)
note(ws, R, 6, "Unclaimed charms use zero storage")

R += 1  # 35 — Re-upload rate during settling
sc(ws, R, 2, "Re-upload rate (% charms re-uploaded in 14-day settling)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.20, pct_fmt)
note(ws, R, 6, "Each re-upload = delete old + write new (extra ops)")

# ════════════════════════════════════════════════════════════════════════════
# SECTION: PLAYBACK & REQUEST VOLUME  (rows 37-43)
# ════════════════════════════════════════════════════════════════════════════
R = 37
sc(ws, R, 2, "PLAYBACK & API REQUEST VOLUME", font=section_font)

R += 1  # 38
sc(ws, R, 2, "Avg views per charm per month (first 3 months)", font=normal_font, border=thin_border)
inp(ws, R, 3, 8, num_fmt)
note(ws, R, 6, "Novelty period — owner + shared views")

R += 1  # 39
sc(ws, R, 2, "Avg views per charm per month (after 3 months)", font=normal_font, border=thin_border)
inp(ws, R, 3, 2, num_fmt)
note(ws, R, 6, "Long-tail: occasional revisits")

R += 1  # 40
sc(ws, R, 2, "% of views requiring glyph verification", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.40, pct_fmt)
note(ws, R, 6, "Each glyph verify = 1 extra API call + Table write")

R += 1  # 41 — API calls per charm lifecycle
sc(ws, R, 2, "API CALLS PER CHARM LIFECYCLE (one-time)", font=bold_font, border=thin_border)
note(ws, R, 6, "claim + configure + get-upload-urls + finalize + owner-preview")
inp(ws, R, 3, 5, num_fmt)

R += 1  # 42 — Table writes per charm lifecycle
sc(ws, R, 2, "Table Storage writes per charm setup", font=normal_font, border=thin_border)
inp(ws, R, 3, 8, num_fmt)
note(ws, R, 6, "charm entity + user-charm + profile + request log (sampled)")

R += 1  # 43 — Table reads per view
sc(ws, R, 2, "Table Storage reads per charm view", font=normal_font, border=thin_border)
inp(ws, R, 3, 3, num_fmt)
note(ws, R, 6, "GetCharm (entity lookup + rate-limit check + request log)")

# ════════════════════════════════════════════════════════════════════════════
# SECTION: AZURE BLOB COOL STORAGE PRICING  (rows 45-51)
# ════════════════════════════════════════════════════════════════════════════
R = 45
sc(ws, R, 2, "AZURE BLOB STORAGE — COOL TIER (Backup)", font=section_font)
note(ws, R, 6, "Azure is backup; R2 is primary CDN delivery")

R += 1  # 46
section_header(ws, R, 2, 4, ["Cost Component", "Rate", "Unit"])

R += 1  # 47
sc(ws, R, 2, "Storage (per GB per month)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.01, currency_micro)
note(ws, R, 4, "$/GB/month")
note(ws, R, 6, "Cool tier — $0.01/GB. Cold = $0.0036 (higher access)")

R += 1  # 48
sc(ws, R, 2, "Write operations (per 10,000 ops)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.10, currency_micro)
note(ws, R, 4, "$/10K writes")
note(ws, R, 6, "PUT/Create blob — initial upload + re-uploads")

R += 1  # 49
sc(ws, R, 2, "Read operations (per 10,000 ops)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.01, currency_micro)
note(ws, R, 4, "$/10K reads")
note(ws, R, 6, "GET blob — fallback downloads only (R2 is primary)")

R += 1  # 50
sc(ws, R, 2, "Data retrieval (per GB downloaded)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.01, currency_micro)
note(ws, R, 4, "$/GB retrieval")
note(ws, R, 6, "Cool tier read penalty — only on Azure fallback reads")

R += 1  # 51
sc(ws, R, 2, "Data egress (per GB, first 100 GB free/mo)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.087, currency_micro)
note(ws, R, 4, "$/GB egress")
note(ws, R, 6, "Only if client falls back from R2 — normally $0")

# ════════════════════════════════════════════════════════════════════════════
# SECTION: CLOUDFLARE R2 PRICING  (rows 53-59)
# ════════════════════════════════════════════════════════════════════════════
R = 53
sc(ws, R, 2, "CLOUDFLARE R2 (Primary CDN — Zero Egress)", font=section_font)
note(ws, R, 6, "S3-compatible; zero bandwidth/egress fees")

R += 1  # 54
section_header(ws, R, 2, 4, ["Cost Component", "Rate", "Unit"])

R += 1  # 55
sc(ws, R, 2, "Storage (per GB per month)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.015, currency_micro)
note(ws, R, 4, "$/GB/month")
note(ws, R, 6, "First 10 GB free; $0.015/GB after")

R += 1  # 56
sc(ws, R, 2, "Class A ops — writes (per 1,000,000 ops)", font=normal_font, border=thin_border)
inp(ws, R, 3, 4.50, currency_fmt)
note(ws, R, 4, "$/1M Class A")
note(ws, R, 6, "PUT/POST/LIST — uploads, re-uploads, finalize checks")

R += 1  # 57
sc(ws, R, 2, "Class B ops — reads (per 1,000,000 ops)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.36, currency_fmt)
note(ws, R, 4, "$/1M Class B")
note(ws, R, 6, "GET — every charm playback downloads from R2")

R += 1  # 58
sc(ws, R, 2, "Egress (bandwidth)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.00, currency_micro)
note(ws, R, 4, "$/GB")
note(ws, R, 6, "ZERO — Cloudflare's key advantage")

R += 1  # 59
sc(ws, R, 2, "R2 free tier (storage GB included free)", font=normal_font, border=thin_border)
inp(ws, R, 3, 10, num_fmt)
note(ws, R, 4, "GB free/month")
note(ws, R, 6, "Also: 10M Class A free, 10M Class B free per month")

# ════════════════════════════════════════════════════════════════════════════
# SECTION: AZURE TABLE STORAGE PRICING  (rows 61-65)
# ════════════════════════════════════════════════════════════════════════════
R = 61
sc(ws, R, 2, "AZURE TABLE STORAGE (Metadata & Telemetry)", font=section_font)

R += 1  # 62
section_header(ws, R, 2, 4, ["Cost Component", "Rate", "Unit"])

R += 1  # 63
sc(ws, R, 2, "Storage (per GB per month)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.045, currency_micro)
note(ws, R, 4, "$/GB/month")
note(ws, R, 6, "Charm entities, profiles, user-charms, request logs")

R += 1  # 64
sc(ws, R, 2, "Transactions (per 10,000)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.00036, currency_micro)
note(ws, R, 4, "$/10K txns")
note(ws, R, 6, "Each read/write/query = 1 transaction")

R += 1  # 65
sc(ws, R, 2, "Avg entity size (KB per charm record set)", font=normal_font, border=thin_border)
inp(ws, R, 3, 2, num_fmt)
note(ws, R, 4, "KB")
note(ws, R, 6, "charm + user-charm + profile share ≈ 2 KB total")

# ════════════════════════════════════════════════════════════════════════════
# SECTION: AZURE FUNCTIONS PRICING  (rows 67-72)
# ════════════════════════════════════════════════════════════════════════════
R = 67
sc(ws, R, 2, "AZURE FUNCTIONS — CONSUMPTION PLAN", font=section_font)

R += 1  # 68
section_header(ws, R, 2, 4, ["Cost Component", "Rate", "Unit"])

R += 1  # 69
sc(ws, R, 2, "Executions (per 1,000,000)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.20, currency_fmt)
note(ws, R, 4, "$/1M executions")
note(ws, R, 6, "First 1M/month FREE; then $0.20/million")

R += 1  # 70
sc(ws, R, 2, "Compute (per GB-second)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.000016, '$#,##0.000000')
note(ws, R, 4, "$/GB-s")
note(ws, R, 6, "First 400K GB-s FREE; ~128MB × 200ms avg per call")

R += 1  # 71
sc(ws, R, 2, "Avg execution duration (ms)", font=normal_font, border=thin_border)
inp(ws, R, 3, 200, num_fmt)
note(ws, R, 4, "milliseconds")
note(ws, R, 6, "Typical for table lookup + SAS URL generation")

R += 1  # 72
sc(ws, R, 2, "Memory allocation (MB)", font=normal_font, border=thin_border)
inp(ws, R, 3, 128, num_fmt)
note(ws, R, 4, "MB")
note(ws, R, 6, "Consumption plan default; 128 MB minimum")

R += 1  # 73
sc(ws, R, 2, "Free tier executions (per month)", font=normal_font, border=thin_border)
inp(ws, R, 3, 1000000, num_fmt)
note(ws, R, 4, "executions")

R += 1  # 74
sc(ws, R, 2, "Free tier GB-seconds (per month)", font=normal_font, border=thin_border)
inp(ws, R, 3, 400000, num_fmt)
note(ws, R, 4, "GB-seconds")

# ════════════════════════════════════════════════════════════════════════════
# SECTION: OTHER PLATFORM COSTS  (rows 76-81)
# ════════════════════════════════════════════════════════════════════════════
R = 76
sc(ws, R, 2, "OTHER PLATFORM COSTS (Monthly Fixed)", font=section_font)

R += 1  # 77
section_header(ws, R, 2, 4, ["Item", "Monthly Cost", "Notes"])

R += 1  # 78
sc(ws, R, 2, "Entra CIAM (Authentication)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.00, currency_fmt)
note(ws, R, 4, "Free: 50K MAU. $0.0025/MAU after that")
note(ws, R, 6, "Per-MAU charge applies only beyond 50K monthly active users")

R += 1  # 79
sc(ws, R, 2, "CIAM per-MAU cost (beyond 50K free)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.0025, currency_micro)
note(ws, R, 4, "$/MAU above 50K")

R += 1  # 80
sc(ws, R, 2, "Domain / SSL / DNS", font=normal_font, border=thin_border)
inp(ws, R, 3, 15.00, currency_fmt)

R += 1  # 81
sc(ws, R, 2, "Application Insights / Monitoring", font=normal_font, border=thin_border)
inp(ws, R, 3, 10.00, currency_fmt)
note(ws, R, 4, "First 5 GB/mo free ingestion")

# ════════════════════════════════════════════════════════════════════════════
# SECTION: OPERATING EXPENSES  (rows 83-89)
# ════════════════════════════════════════════════════════════════════════════
R = 83
sc(ws, R, 2, "OPERATING EXPENSES (Monthly)", font=section_font)
R += 1  # 84
section_header(ws, R, 2, 4, ["Expense", "Monthly Cost", "MoM Growth %"])

R += 1  # 85
sc(ws, R, 2, "Marketing / Customer Acquisition", font=normal_font, border=thin_border)
inp(ws, R, 3, 500, currency_whole); inp(ws, R, 4, 0.05, pct_fmt)

R += 1  # 86
sc(ws, R, 2, "Shipping & Fulfillment (per charm shipped)", font=normal_font, border=thin_border)
inp(ws, R, 3, 3.50, currency_fmt)
note(ws, R, 4, "Per unit shipped")

R += 1  # 87
sc(ws, R, 2, "Payment Processing (% of revenue)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.029, pct_fmt)
note(ws, R, 4, "Stripe / processor fee")

R += 1  # 88
sc(ws, R, 2, "Customer Support", font=normal_font, border=thin_border)
inp(ws, R, 3, 200, currency_whole); inp(ws, R, 4, 0.03, pct_fmt)

R += 1  # 89
sc(ws, R, 2, "Insurance / Legal / Misc", font=normal_font, border=thin_border)
inp(ws, R, 3, 150, currency_whole); inp(ws, R, 4, 0.00, pct_fmt)

# ════════════════════════════════════════════════════════════════════════════
# SECTION: TAX  (row 91)
# ════════════════════════════════════════════════════════════════════════════
R = 91
sc(ws, R, 2, "OTHER ASSUMPTIONS", font=section_font)
R += 1  # 92
sc(ws, R, 2, "Tax rate (on positive EBITDA)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.21, pct_fmt)

# ════════════════════════════════════════════════════════════════════════════
# SECTION: RETURNS & REPLACEMENTS  (rows 94-103)
# ════════════════════════════════════════════════════════════════════════════
R = 94
sc(ws, R, 2, "RETURNS & REPLACEMENTS", font=section_font)
note(ws, R, 6, "Industry avg for specialty/gift items: 5-10% returns, 2-4% defects")

R += 1  # 95
section_header(ws, R, 2, 4, ["Item", "Rate / Cost", "Notes"])

R += 1  # 96
sc(ws, R, 2, "Return rate (% of monthly charm sales returned)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.06, pct_fmt)
note(ws, R, 4, "Specialty/gift items — lower than general e-commerce")
note(ws, R, 6, "E-commerce avg ~15-30%; gift/keepsake items typically 5-8%")

R += 1  # 97
sc(ws, R, 2, "% of returns that are pre-claim (restockable)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.60, pct_fmt)
note(ws, R, 4, "Charm bought but never claimed/configured")
note(ws, R, 6, "Pre-claim charms can be wiped and resold; post-claim are dead stock")

R += 1  # 98
sc(ws, R, 2, "Restocking salvage (% of COGS recovered on pre-claim returns)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.80, pct_fmt)
note(ws, R, 4, "Repackage & resell; some cosmetic loss")
note(ws, R, 6, "Pre-claim charms only — post-claim have content, cannot resell")

R += 1  # 99
sc(ws, R, 2, "Return shipping cost (company pays, per unit)", font=normal_font, border=thin_border)
inp(ws, R, 3, 5.00, currency_fmt)
note(ws, R, 6, "Return label cost; higher than outbound due to reverse logistics")

R += 1  # 100
sc(ws, R, 2, "Return processing / handling cost (per unit)", font=normal_font, border=thin_border)
inp(ws, R, 3, 2.00, currency_fmt)
note(ws, R, 6, "Inspect, wipe NFC, repackage or dispose")

R += 1  # 101
sc(ws, R, 2, "Replacement / defect rate (% of sales needing free replacement)", font=normal_font, border=thin_border)
inp(ws, R, 3, 0.03, pct_fmt)
note(ws, R, 4, "NFC failure, shipping damage, manufacturing defect")
note(ws, R, 6, "Industry avg for electronics-adjacent: 2-5%")

R += 1  # 102
sc(ws, R, 2, "Replacement shipping cost (per unit)", font=normal_font, border=thin_border)
inp(ws, R, 3, 3.50, currency_fmt)
note(ws, R, 6, "Same as outbound; may include expedited for goodwill")

R += 1  # 103
sc(ws, R, 2, "Replacement COGS (% of original — may use cheaper housing)", font=normal_font, border=thin_border)
inp(ws, R, 3, 1.00, pct_fmt)
note(ws, R, 4, "100% = full COGS; could be lower if reusing parts")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: 24-MONTH PROJECTIONS
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("24-Month Projections")
ws2.sheet_properties.tabColor = "2E7D32"
ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 48

sc(ws2, 1, 2, "MemoryCharm — 24-Month Financial Projections", font=title_font)
sc(ws2, 2, 2, "All values driven by Assumptions sheet — change inputs there", font=Font(name="Calibri", italic=True, size=10, color="666666"))

# Month header row = 4
R = 4
sc(ws2, R, 2, "", font=header_font, fill=header_fill)
for m in range(1, 25):
    col = m + 2
    ws2.column_dimensions[get_column_letter(col)].width = 14
    sc(ws2, R, col, f"Month {m}", font=header_font, fill=header_fill, alignment=Alignment(horizontal="center"))

ASM = "Assumptions"  # Sheet reference shorthand


def proj_row(ws, row, label, font_=None, fmt=None, fill_=None):
    """Label a projection row."""
    sc(ws, row, 2, label, font=font_ or normal_font, border=thin_border)
    if fill_:
        style_range(ws, row, 2, 26, fill=fill_)


def proj_formula(ws, row, month_formulas_fn, fmt=None, font_=None, fill_=None):
    """Fill months 1-24 with formulas returned by month_formulas_fn(m, col, col_letter)."""
    for m in range(1, 25):
        col = m + 2
        c = get_column_letter(col)
        prev = get_column_letter(col - 1) if m > 1 else None
        f = month_formulas_fn(m, col, c, prev)
        cell = ws.cell(row=row, column=col)
        cell.value = f
        cell.font = font_ or normal_font
        cell.number_format = fmt or num_fmt
        cell.border = thin_border
        if fill_:
            cell.fill = fill_


# ── UNIT SALES (rows 5-10) ──────────────────────────────────────────────────
R = 5; sc(ws2, R, 2, "UNIT SALES", font=section_font)

# Row 6: 10-Year sold
R = 6; proj_row(ws2, R, "10-Year Charms Sold")
proj_formula(ws2, R, lambda m,col,c,p: f"={ASM}!C18" if m==1 else f"=ROUND({p}{R}*(1+{ASM}!D18),0)")

# Row 7: 15-Year sold
R = 7; proj_row(ws2, R, "15-Year Charms Sold")
proj_formula(ws2, R, lambda m,col,c,p: f"={ASM}!C19" if m==1 else f"=ROUND({p}{R}*(1+{ASM}!D19),0)")

# Row 8: Perpetual sold
R = 8; proj_row(ws2, R, "Retail (Perpetual) Charms Sold")
proj_formula(ws2, R, lambda m,col,c,p: f"={ASM}!C20" if m==1 else f"=ROUND({p}{R}*(1+{ASM}!D20),0)")

# Row 9: Total monthly
R = 9; proj_row(ws2, R, "Total Charms Sold (Month)", bold_font, fill_=blue_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"={c}6+{c}7+{c}8", font_=bold_font, fill_=blue_fill)

# Row 10: Cumulative
R = 10; proj_row(ws2, R, "Cumulative Charms Sold", bold_font, fill_=blue_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"={c}9" if m==1 else f"={p}{R}+{c}9", font_=bold_font, fill_=blue_fill)

# Row 11: Claimed (active) charms cumulative
R = 11; proj_row(ws2, R, "Active Charms (claimed, storing content)")
proj_formula(ws2, R, lambda m,col,c,p: f"=ROUND({c}10*{ASM}!C34,0)")

# ── STORAGE VOLUME (rows 13-17) ─────────────────────────────────────────────
R = 13; sc(ws2, R, 2, "STORAGE VOLUME", font=section_font)

# Row 14: New content uploaded this month (GB)
R = 14; proj_row(ws2, R, "New Content Uploaded (GB)")
proj_formula(ws2, R,
    lambda m,col,c,p: f"=ROUND({c}9*{ASM}!C34*{ASM}!C33/1024,2)",
    fmt=num_2dp)
note(ws2, R, 2, None)  # label already set

# Row 15: Cumulative stored (GB) — both Azure + R2
R = 15; proj_row(ws2, R, "Cumulative Content Stored (GB)", bold_font)
proj_formula(ws2, R,
    lambda m,col,c,p: f"={c}14" if m==1 else f"={p}{R}+{c}14",
    fmt=num_1dp, font_=bold_font)

# Row 16: R2 storage (same as cumulative, since R2 is primary)
R = 16; proj_row(ws2, R, "  Cloudflare R2 Storage (GB)")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}15", fmt=num_1dp)

# Row 17: Azure Blob Cool storage (same — it's the backup mirror)
R = 17; proj_row(ws2, R, "  Azure Blob Cool Storage (GB)")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}15", fmt=num_1dp)

# ── REQUEST VOLUME (rows 19-27) ─────────────────────────────────────────────
R = 19; sc(ws2, R, 2, "API & STORAGE REQUEST VOLUME", font=section_font)

# Row 20: Total monthly playback views
# Charms sold in last 3 months get "new" view rate; older get "long-tail" rate
R = 20; proj_row(ws2, R, "Total Charm Playback Views (month)")
# Simplified: new charms (this month's sales) * high rate + older cumulative * low rate
for m in range(1, 25):
    col = m + 2
    c = get_column_letter(col)
    # Charms sold in last 3 months = sum of last 3 months of sales (or fewer if m<3)
    if m == 1:
        new_charms = f"{c}9"
    elif m == 2:
        new_charms = f"({c}9+{get_column_letter(col-1)}9)"
    else:
        new_charms = f"({c}9+{get_column_letter(col-1)}9+{get_column_letter(col-2)}9)"
    old_charms = f"MAX(0,{c}11-{new_charms})"
    f = f"=ROUND({new_charms}*{ASM}!C38+{old_charms}*{ASM}!C39,0)"
    cell = ws2.cell(row=R, column=col)
    cell.value = f; cell.font = normal_font; cell.number_format = num_fmt; cell.border = thin_border

# Row 21: Glyph verification calls
R = 21; proj_row(ws2, R, "Glyph Verification API Calls")
proj_formula(ws2, R, lambda m,col,c,p: f"=ROUND({c}20*{ASM}!C40,0)")

# Row 22: Total Azure Functions invocations
R = 22; proj_row(ws2, R, "Total Azure Functions Invocations", bold_font)
# = (new charms * lifecycle calls) + (views * 1 GetCharm call each) + (glyph calls) + (admin overhead ~2%)
proj_formula(ws2, R,
    lambda m,col,c,p: f"=ROUND(({c}9*{ASM}!C34*{ASM}!C41)+{c}20+{c}21+({c}20*0.02),0)",
    font_=bold_font)

# Row 23: R2 write operations (Class A) — initial upload + re-uploads
R = 23; proj_row(ws2, R, "R2 Class A Ops (Writes)")
# Each claimed charm = ~2 writes (PUT blob + list verify). Re-uploads add more.
proj_formula(ws2, R,
    lambda m,col,c,p: f"=ROUND({c}9*{ASM}!C34*2*(1+{ASM}!C35),0)")

# Row 24: R2 read operations (Class B) — every playback view
R = 24; proj_row(ws2, R, "R2 Class B Ops (Reads / Playback)")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}20")

# Row 25: Azure Blob write operations (mirror of R2 writes)
R = 25; proj_row(ws2, R, "Azure Blob Write Ops (Backup Uploads)")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}23")

# Row 26: Azure Blob read operations (fallback only — small %)
R = 26; proj_row(ws2, R, "Azure Blob Read Ops (Fallback — est 2%)")
proj_formula(ws2, R, lambda m,col,c,p: f"=ROUND({c}20*0.02,0)")

# Row 27: Table Storage transactions
R = 27; proj_row(ws2, R, "Azure Table Transactions (Total)", bold_font)
# = (new charms * setup writes) + (all views * reads per view) + (glyph * 2 writes each)
proj_formula(ws2, R,
    lambda m,col,c,p: f"=ROUND({c}9*{ASM}!C34*{ASM}!C42+{c}20*{ASM}!C43+{c}21*2,0)",
    font_=bold_font)

# Row 28: Playback bandwidth from R2 (GB)
R = 28; proj_row(ws2, R, "R2 Playback Bandwidth (GB) — FREE egress")
proj_formula(ws2, R,
    lambda m,col,c,p: f"=ROUND({c}20*{ASM}!C33/1024,1)",
    fmt=num_1dp)

# ── INFRASTRUCTURE COSTS (rows 30-42) ───────────────────────────────────────
R = 30; sc(ws2, R, 2, "INFRASTRUCTURE COST BREAKDOWN", font=section_font)

# Row 31: Azure Blob Cool — Storage
R = 31; proj_row(ws2, R, "  Azure Blob: Storage Cost")
proj_formula(ws2, R,
    lambda m,col,c,p: f"={c}17*{ASM}!C47",
    fmt=currency_fmt)

# Row 32: Azure Blob Cool — Write ops
R = 32; proj_row(ws2, R, "  Azure Blob: Write Operations")
proj_formula(ws2, R,
    lambda m,col,c,p: f"=({c}25/10000)*{ASM}!C48",
    fmt=currency_fmt)

# Row 33: Azure Blob Cool — Read ops (fallback)
R = 33; proj_row(ws2, R, "  Azure Blob: Read Ops (Fallback)")
proj_formula(ws2, R,
    lambda m,col,c,p: f"=({c}26/10000)*{ASM}!C49",
    fmt=currency_fmt)

# Row 34: Azure Blob — Data retrieval (fallback reads)
R = 34; proj_row(ws2, R, "  Azure Blob: Data Retrieval (Fallback)")
proj_formula(ws2, R,
    lambda m,col,c,p: f"=({c}26*{ASM}!C33/1024)*{ASM}!C50",
    fmt=currency_fmt)

# Row 35: TOTAL Azure Blob
R = 35; proj_row(ws2, R, "Azure Blob Total", bold_font)
proj_formula(ws2, R,
    lambda m,col,c,p: f"={c}31+{c}32+{c}33+{c}34",
    fmt=currency_fmt, font_=bold_font)

# Row 36: R2 — Storage cost (minus free tier)
R = 36; proj_row(ws2, R, "  R2: Storage Cost")
proj_formula(ws2, R,
    lambda m,col,c,p: f"=MAX(0,{c}16-{ASM}!C59)*{ASM}!C55",
    fmt=currency_fmt)

# Row 37: R2 — Class A (writes)
R = 37; proj_row(ws2, R, "  R2: Class A Ops (Writes)")
proj_formula(ws2, R,
    lambda m,col,c,p: f"=MAX(0,{c}23-10000000)*{ASM}!C56/1000000",
    fmt=currency_fmt)

# Row 38: R2 — Class B (reads/playback)
R = 38; proj_row(ws2, R, "  R2: Class B Ops (Reads)")
proj_formula(ws2, R,
    lambda m,col,c,p: f"=MAX(0,{c}24-10000000)*{ASM}!C57/1000000",
    fmt=currency_fmt)

# Row 39: TOTAL R2
R = 39; proj_row(ws2, R, "Cloudflare R2 Total", bold_font)
proj_formula(ws2, R,
    lambda m,col,c,p: f"={c}36+{c}37+{c}38",
    fmt=currency_fmt, font_=bold_font)

# Row 40: Azure Table Storage
R = 40; proj_row(ws2, R, "Azure Table Storage")
# Storage cost + transaction cost
proj_formula(ws2, R,
    lambda m,col,c,p: f"=({c}11*{ASM}!C65/1024/1024)*{ASM}!C63+({c}27/10000)*{ASM}!C64",
    fmt=currency_fmt)

# Row 41: Azure Functions
R = 41; proj_row(ws2, R, "Azure Functions (Compute)")
# Execution cost (above free tier) + GB-seconds cost (above free tier)
proj_formula(ws2, R,
    lambda m,col,c,p: (
        f"=MAX(0,{c}22-{ASM}!C73)/1000000*{ASM}!C69"
        f"+MAX(0,{c}22*({ASM}!C71/1000)*({ASM}!C72/1024)-{ASM}!C74)*{ASM}!C70"
    ),
    fmt=currency_fmt)

# Row 42: Other platform (CIAM + DNS + Monitoring)
R = 42; proj_row(ws2, R, "Other Platform (CIAM + DNS + Monitoring)")
# CIAM: free up to 50K MAU, then per-MAU. Approximate MAU ≈ active charms * 0.3 (not every charm = unique user)
proj_formula(ws2, R,
    lambda m,col,c,p: f"={ASM}!C78+MAX(0,{c}11*0.3-50000)*{ASM}!C79+{ASM}!C80+{ASM}!C81",
    fmt=currency_fmt)

# Row 43: TOTAL INFRASTRUCTURE
R = 43; proj_row(ws2, R, "TOTAL INFRASTRUCTURE", bold_font, fill_=red_fill)
proj_formula(ws2, R,
    lambda m,col,c,p: f"={c}35+{c}39+{c}40+{c}41+{c}42",
    fmt=currency_fmt, font_=bold_font, fill_=red_fill)

# ── REVENUE (rows 45-52) ────────────────────────────────────────────────────
R = 45; sc(ws2, R, 2, "REVENUE", font=section_font)

R = 46; proj_row(ws2, R, "10-Year Charm Revenue")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}6*{ASM}!C6", fmt=currency_whole)

R = 47; proj_row(ws2, R, "15-Year Charm Revenue")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}7*{ASM}!C7", fmt=currency_whole)

R = 48; proj_row(ws2, R, "Retail (Perpetual) Charm Revenue")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}8*{ASM}!C8", fmt=currency_whole)

R = 49; proj_row(ws2, R, "Extend Memory Revenue")
proj_formula(ws2, R, lambda m,col,c,p: f"=ROUND({c}10*{ASM}!C23,0)*{ASM}!C12", fmt=currency_whole)

R = 50; proj_row(ws2, R, "Upgrade Tier Revenue")
proj_formula(ws2, R, lambda m,col,c,p: f"=ROUND({c}10*{ASM}!C24,0)*{ASM}!C13", fmt=currency_whole)

R = 51; proj_row(ws2, R, "Gift Wrap Revenue")
proj_formula(ws2, R, lambda m,col,c,p: f"=ROUND({c}9*{ASM}!C25,0)*{ASM}!C14", fmt=currency_whole)

R = 52; proj_row(ws2, R, "TOTAL REVENUE", bold_font, fill_=green_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"=SUM({c}46:{c}51)", fmt=currency_whole, font_=bold_font, fill_=green_fill)

# ── COGS (rows 54-59) ───────────────────────────────────────────────────────
R = 54; sc(ws2, R, 2, "COST OF GOODS SOLD", font=section_font)

R = 55; proj_row(ws2, R, "10-Year Charm COGS")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}6*{ASM}!D6", fmt=currency_whole)

R = 56; proj_row(ws2, R, "15-Year Charm COGS")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}7*{ASM}!D7", fmt=currency_whole)

R = 57; proj_row(ws2, R, "Retail (Perpetual) Charm COGS")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}8*{ASM}!D8", fmt=currency_whole)

R = 58; proj_row(ws2, R, "Upsell COGS")
proj_formula(ws2, R,
    lambda m,col,c,p: (f"=ROUND({c}10*{ASM}!C23,0)*{ASM}!D12"
                       f"+ROUND({c}10*{ASM}!C24,0)*{ASM}!D13"
                       f"+ROUND({c}9*{ASM}!C25,0)*{ASM}!D14"),
    fmt=currency_whole)

R = 59; proj_row(ws2, R, "TOTAL COGS", bold_font, fill_=red_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"=SUM({c}55:{c}58)", fmt=currency_whole, font_=bold_font, fill_=red_fill)

# ── GROSS PROFIT (row 60-61) ────────────────────────────────────────────────
R = 60; proj_row(ws2, R, "GROSS PROFIT", bold_font, fill_=green_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"={c}52-{c}59", fmt=currency_whole, font_=bold_font, fill_=green_fill)

R = 61; proj_row(ws2, R, "Gross Margin %", bold_font)
proj_formula(ws2, R, lambda m,col,c,p: f"=IF({c}52>0,{c}60/{c}52,0)", fmt=pct_fmt, font_=bold_font)

# ── OPERATING EXPENSES (rows 63-70) ─────────────────────────────────────────
R = 63; sc(ws2, R, 2, "OPERATING EXPENSES", font=section_font)

R = 64; proj_row(ws2, R, "Infrastructure (see breakdown above)")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}43", fmt=currency_whole)

R = 65; proj_row(ws2, R, "Marketing / Customer Acquisition")
proj_formula(ws2, R,
    lambda m,col,c,p: f"={ASM}!C85" if m==1 else f"=ROUND({p}{R}*(1+{ASM}!D85),0)",
    fmt=currency_whole)

R = 66; proj_row(ws2, R, "Shipping & Fulfillment")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}9*{ASM}!C86", fmt=currency_whole)

R = 67; proj_row(ws2, R, "Payment Processing")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}52*{ASM}!C87", fmt=currency_whole)

R = 68; proj_row(ws2, R, "Customer Support")
proj_formula(ws2, R,
    lambda m,col,c,p: f"={ASM}!C88" if m==1 else f"=ROUND({p}{R}*(1+{ASM}!D88),0)",
    fmt=currency_whole)

R = 69; proj_row(ws2, R, "Insurance / Legal / Misc")
proj_formula(ws2, R,
    lambda m,col,c,p: f"={ASM}!C89" if m==1 else f"=ROUND({p}{R}*(1+{ASM}!D89),0)",
    fmt=currency_whole)

R = 70; proj_row(ws2, R, "TOTAL OPERATING EXPENSES", bold_font, fill_=red_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"=SUM({c}64:{c}69)", fmt=currency_whole, font_=bold_font, fill_=red_fill)

# ── P&L (rows 72-79) ────────────────────────────────────────────────────────
R = 72; sc(ws2, R, 2, "PROFIT & LOSS", font=section_font)

R = 73; proj_row(ws2, R, "EBITDA (Gross Profit - OpEx)", bold_font, fill_=green_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"={c}60-{c}70", fmt=currency_whole, font_=bold_font, fill_=green_fill)

R = 74; proj_row(ws2, R, "EBITDA Margin %", bold_font)
proj_formula(ws2, R, lambda m,col,c,p: f"=IF({c}52>0,{c}73/{c}52,0)", fmt=pct_fmt, font_=bold_font)

R = 75; proj_row(ws2, R, "Tax (on positive EBITDA)")
proj_formula(ws2, R, lambda m,col,c,p: f"=IF({c}73>0,{c}73*{ASM}!C92,0)", fmt=currency_whole)

big_font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)
R = 76; proj_row(ws2, R, "NET INCOME", big_font, fill_=green_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"={c}73-{c}75", fmt=currency_whole, font_=big_font, fill_=green_fill)

R = 77; proj_row(ws2, R, "Cumulative Net Income", bold_font)
proj_formula(ws2, R,
    lambda m,col,c,p: f"={c}76" if m==1 else f"={p}{R}+{c}76",
    fmt=currency_whole, font_=bold_font)

# ── KEY METRICS (rows 79-85) ────────────────────────────────────────────────
R = 79; sc(ws2, R, 2, "KEY METRICS", font=section_font)

R = 80; proj_row(ws2, R, "Customer Acquisition Cost (CAC)")
proj_formula(ws2, R, lambda m,col,c,p: f"=IF({c}9>0,{c}65/{c}9,0)", fmt=currency_fmt)

R = 81; proj_row(ws2, R, "Avg Revenue Per Charm Sold")
proj_formula(ws2, R, lambda m,col,c,p: f"=IF({c}9>0,{c}52/{c}9,0)", fmt=currency_fmt)

R = 82; proj_row(ws2, R, "All-in Cost Per Charm (COGS+OpEx)")
proj_formula(ws2, R, lambda m,col,c,p: f"=IF({c}9>0,({c}59+{c}70)/{c}9,0)", fmt=currency_fmt)

R = 83; proj_row(ws2, R, "Infra Cost Per Active Charm (Monthly)")
proj_formula(ws2, R, lambda m,col,c,p: f"=IF({c}11>0,{c}43/{c}11,0)", fmt=currency_micro)

R = 84; proj_row(ws2, R, "Total Storage (GB — Azure + R2 combined)")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}15*2", fmt=num_1dp)

R = 85; proj_row(ws2, R, 'Monthly Profitable? (EBITDA > 0)')
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws2.cell(row=85, column=col)
    cell.value = f'=IF({c}73>0,"YES","NO")'
    cell.font = bold_font; cell.border = thin_border; cell.alignment = Alignment(horizontal="center")

# ── RETURNS & REPLACEMENTS (rows 87-101) ─────────────────────────────────────
R = 87; sc(ws2, R, 2, "RETURNS & REPLACEMENTS", font=section_font)

# Row 88: Returned units
R = 88; proj_row(ws2, R, "Returned Charms (units)")
proj_formula(ws2, R, lambda m,col,c,p: f"=ROUND({c}9*{ASM}!C96,0)")

# Row 89: Of which pre-claim (restockable)
R = 89; proj_row(ws2, R, "  Pre-Claim Returns (restockable)")
proj_formula(ws2, R, lambda m,col,c,p: f"=ROUND({c}88*{ASM}!C97,0)")

# Row 90: Of which post-claim (dead stock)
R = 90; proj_row(ws2, R, "  Post-Claim Returns (dead stock)")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}88-{c}89")

# Row 91: Replacement units (defects)
R = 91; proj_row(ws2, R, "Replacement Charms Shipped (defects)")
proj_formula(ws2, R, lambda m,col,c,p: f"=ROUND({c}9*{ASM}!C101,0)")

# Row 92: blank separator
R = 93; proj_row(ws2, R, "RETURN & REPLACEMENT COSTS", bold_font)

# Row 94: Refund amount (returned units * avg weighted price)
R = 94; proj_row(ws2, R, "Refund Amount (returned units x avg price)")
# Use weighted avg price: total revenue / total charms sold
proj_formula(ws2, R,
    lambda m,col,c,p: f"=IF({c}9>0, {c}88*({c}52/{c}9), 0)",
    fmt=currency_whole)

# Row 95: Return shipping cost
R = 95; proj_row(ws2, R, "Return Shipping Cost")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}88*{ASM}!C99", fmt=currency_whole)

# Row 96: Return processing cost
R = 96; proj_row(ws2, R, "Return Processing / Handling")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}88*{ASM}!C100", fmt=currency_whole)

# Row 97: COGS lost on post-claim returns (non-restockable — full loss)
R = 97; proj_row(ws2, R, "COGS Lost — Post-Claim Returns (dead stock)")
# Post-claim charms: COGS not recoverable. Use avg COGS = total COGS / total units
proj_formula(ws2, R,
    lambda m,col,c,p: f"=IF({c}9>0, {c}90*({c}59/{c}9), 0)",
    fmt=currency_whole)

# Row 98: COGS salvage on pre-claim returns
R = 98; proj_row(ws2, R, "COGS Salvaged — Pre-Claim Returns (restocked)")
proj_formula(ws2, R,
    lambda m,col,c,p: f"=IF({c}9>0, -{c}89*({c}59/{c}9)*{ASM}!C98, 0)",
    fmt=currency_whole)

# Row 99: Replacement COGS (defect units * COGS * replacement rate)
R = 99; proj_row(ws2, R, "Replacement COGS (defect units x COGS)")
proj_formula(ws2, R,
    lambda m,col,c,p: f"=IF({c}9>0, {c}91*({c}59/{c}9)*{ASM}!C103, 0)",
    fmt=currency_whole)

# Row 100: Replacement shipping
R = 100; proj_row(ws2, R, "Replacement Shipping")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}91*{ASM}!C102", fmt=currency_whole)

# Row 101: TOTAL return & replacement cost
R = 101; proj_row(ws2, R, "TOTAL RETURNS & REPLACEMENT IMPACT", bold_font, fill_=red_fill)
# = refund + return shipping + processing + dead stock COGS + salvage offset + replacement COGS + replacement shipping
proj_formula(ws2, R,
    lambda m,col,c,p: f"={c}94+{c}95+{c}96+{c}97+{c}98+{c}99+{c}100",
    fmt=currency_whole, font_=bold_font, fill_=red_fill)

# ── ADJUSTED P&L (rows 103-111) ─────────────────────────────────────────────
R = 103; sc(ws2, R, 2, "ADJUSTED P&L (After Returns & Replacements)", font=section_font)

# Row 104: Net Revenue (gross revenue - refunds)
R = 104; proj_row(ws2, R, "Net Revenue (gross - refunds)", bold_font, fill_=green_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"={c}52-{c}94", fmt=currency_whole, font_=bold_font, fill_=green_fill)

# Row 105: Adjusted COGS (original + dead stock + replacement - salvage)
R = 105; proj_row(ws2, R, "Adjusted COGS (incl. dead stock + replacements)", bold_font, fill_=red_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"={c}59+{c}97+{c}98+{c}99", fmt=currency_whole, font_=bold_font, fill_=red_fill)

# Row 106: Return & replacement overhead (shipping + processing)
R = 106; proj_row(ws2, R, "Return & Replacement Overhead (shipping + processing)")
proj_formula(ws2, R, lambda m,col,c,p: f"={c}95+{c}96+{c}100", fmt=currency_whole)

# Row 107: Adjusted Gross Profit
R = 107; proj_row(ws2, R, "Adjusted Gross Profit", bold_font, fill_=green_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"={c}104-{c}105-{c}106", fmt=currency_whole, font_=bold_font, fill_=green_fill)

# Row 108: Adjusted Gross Margin %
R = 108; proj_row(ws2, R, "Adjusted Gross Margin %", bold_font)
proj_formula(ws2, R, lambda m,col,c,p: f"=IF({c}104>0,{c}107/{c}104,0)", fmt=pct_fmt, font_=bold_font)

# Row 109: Adjusted EBITDA (adjusted gross profit - OpEx)
R = 109; proj_row(ws2, R, "Adjusted EBITDA", bold_font, fill_=green_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"={c}107-{c}70", fmt=currency_whole, font_=bold_font, fill_=green_fill)

# Row 110: Adjusted EBITDA Margin %
R = 110; proj_row(ws2, R, "Adjusted EBITDA Margin %", bold_font)
proj_formula(ws2, R, lambda m,col,c,p: f"=IF({c}104>0,{c}109/{c}104,0)", fmt=pct_fmt, font_=bold_font)

# Row 111: Adjusted Net Income
R = 111; proj_row(ws2, R, "Adjusted Net Income", big_font, fill_=green_fill)
proj_formula(ws2, R, lambda m,col,c,p: f"={c}109-IF({c}109>0,{c}109*{ASM}!C92,0)", fmt=currency_whole, font_=big_font, fill_=green_fill)

# Row 112: Net active charms (adjusted for returns — affects storage)
R = 112; proj_row(ws2, R, "Net Active Charms (adjusted for returns)", bold_font)
# Cumulative sold - cumulative returned (post-claim returns remove content)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    if m == 1:
        f = f"={c}11-{c}90"
    else:
        prev = get_column_letter(col - 1)
        f = f"={prev}{R}+({c}9*{ASM}!C34)-{c}90"
    cell = ws2.cell(row=R, column=col); cell.value = f
    cell.font = bold_font; cell.number_format = num_fmt; cell.border = thin_border

# Row 113: Margin erosion from returns (% of gross revenue lost to returns & replacements)
R = 113; proj_row(ws2, R, "Margin Erosion from Returns (% of gross revenue)")
proj_formula(ws2, R, lambda m,col,c,p: f"=IF({c}52>0,{c}101/{c}52,0)", fmt=pct_fmt)

ws2.freeze_panes = "C5"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: ANNUAL SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Annual Summary")
ws3.sheet_properties.tabColor = ACCENT_GOLD
ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 48
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 18

sc(ws3, 1, 2, "MemoryCharm — Annual Summary", font=title_font)
R = 3
section_header(ws3, R, 2, 5, ["", "Year 1 (M1-12)", "Year 2 (M13-24)", "Total (24 Mo)"])

P = "'24-Month Projections'"

summary_rows = [
    # (row, label, proj_row_or_special, fmt, is_bold, fill)
    (5, "Total Charms Sold", 9, num_fmt, True, None),
    (6, "Cumulative Charms (End)", "cum10", num_fmt, True, None),
    (7, "Active Charms (End)", "cum11", num_fmt, False, None),
    (8, "Total Storage (GB, end of period)", "cum15", num_1dp, False, None),
    (9, "", None, None, False, None),
    (10, "Total Revenue", 52, currency_whole, True, None),
    (11, "  Charm Revenue", "sum46_48", currency_whole, False, None),
    (12, "  Upsell Revenue", "sum49_51", currency_whole, False, None),
    (13, "", None, None, False, None),
    (14, "Total COGS", 59, currency_whole, True, red_fill),
    (15, "Gross Profit", 60, currency_whole, True, green_fill),
    (16, "Gross Margin %", "gm", pct_fmt, True, None),
    (17, "", None, None, False, None),
    (18, "Total Operating Expenses", 70, currency_whole, True, red_fill),
    (19, "  Infrastructure", 43, currency_whole, False, None),
    (20, "    Azure Blob Storage", 35, currency_whole, False, None),
    (21, "    Cloudflare R2", 39, currency_whole, False, None),
    (22, "    Azure Table Storage", 40, currency_whole, False, None),
    (23, "    Azure Functions", 41, currency_whole, False, None),
    (24, "    Other Platform", 42, currency_whole, False, None),
    (25, "  Marketing", 65, currency_whole, False, None),
    (26, "  Shipping", 66, currency_whole, False, None),
    (27, "  Payment Processing", 67, currency_whole, False, None),
    (28, "  Support + Legal", "sum68_69", currency_whole, False, None),
    (29, "", None, None, False, None),
    (30, "EBITDA", 73, currency_whole, True, green_fill),
    (31, "EBITDA Margin %", "em", pct_fmt, True, None),
    (32, "Tax", 75, currency_whole, False, None),
    (33, "NET INCOME", 76, currency_whole, True, green_fill),
    (34, "", None, None, False, None),
    (35, "RETURNS & REPLACEMENTS", None, None, True, None),
    (36, "  Returned Units", 88, num_fmt, False, None),
    (37, "  Replacement Units", 91, num_fmt, False, None),
    (38, "  Total Return & Replacement Cost", 101, currency_whole, True, red_fill),
    (39, "  Margin Erosion %", "me", pct_fmt, False, None),
    (40, "", None, None, False, None),
    (41, "ADJUSTED P&L (After Returns)", None, None, True, None),
    (42, "  Net Revenue", 104, currency_whole, True, None),
    (43, "  Adjusted COGS", 105, currency_whole, False, red_fill),
    (44, "  Adjusted Gross Profit", 107, currency_whole, True, green_fill),
    (45, "  Adjusted Gross Margin %", "agm", pct_fmt, True, None),
    (46, "  Adjusted EBITDA", 109, currency_whole, True, green_fill),
    (47, "  Adjusted EBITDA Margin %", "aem", pct_fmt, True, None),
    (48, "  ADJUSTED NET INCOME", 111, currency_whole, True, green_fill),
]

for (row, label, src, fmt, is_bold, fill) in summary_rows:
    if not label:
        continue
    f = bold_font if is_bold else normal_font
    if label == "NET INCOME":
        f = big_font

    sc(ws3, row, 2, label, font=f, border=thin_border)
    if fill:
        ws3.cell(row=row, column=2).fill = fill

    if isinstance(src, int):
        y1 = f"=SUM({P}!C{src}:N{src})"
        y2 = f"=SUM({P}!O{src}:Z{src})"
        tot = f"=C{row}+D{row}"
    elif src == "cum10":
        y1 = f"={P}!N10"; y2 = f"={P}!Z10"; tot = f"=D{row}"
    elif src == "cum11":
        y1 = f"={P}!N11"; y2 = f"={P}!Z11"; tot = f"=D{row}"
    elif src == "cum15":
        y1 = f"={P}!N15"; y2 = f"={P}!Z15"; tot = f"=D{row}"
    elif src == "sum46_48":
        y1 = f"=SUM({P}!C46:N46)+SUM({P}!C47:N47)+SUM({P}!C48:N48)"
        y2 = f"=SUM({P}!O46:Z46)+SUM({P}!O47:Z47)+SUM({P}!O48:Z48)"
        tot = f"=C{row}+D{row}"
    elif src == "sum49_51":
        y1 = f"=SUM({P}!C49:N49)+SUM({P}!C50:N50)+SUM({P}!C51:N51)"
        y2 = f"=SUM({P}!O49:Z49)+SUM({P}!O50:Z50)+SUM({P}!O51:Z51)"
        tot = f"=C{row}+D{row}"
    elif src == "sum68_69":
        y1 = f"=SUM({P}!C68:N68)+SUM({P}!C69:N69)"
        y2 = f"=SUM({P}!O68:Z68)+SUM({P}!O69:Z69)"
        tot = f"=C{row}+D{row}"
    elif src == "gm":
        y1 = f"=IF(C10>0,C15/C10,0)"
        y2 = f"=IF(D10>0,D15/D10,0)"
        tot = f"=IF(E10>0,E15/E10,0)"
    elif src == "em":
        y1 = f"=IF(C10>0,C30/C10,0)"
        y2 = f"=IF(D10>0,D30/D10,0)"
        tot = f"=IF(E10>0,E30/E10,0)"
    elif src == "me":
        y1 = f"=IF(C10>0,C38/C10,0)"
        y2 = f"=IF(D10>0,D38/D10,0)"
        tot = f"=IF(E10>0,E38/E10,0)"
    elif src == "agm":
        y1 = f"=IF(C42>0,C44/C42,0)"
        y2 = f"=IF(D42>0,D44/D42,0)"
        tot = f"=IF(E42>0,E44/E42,0)"
    elif src == "aem":
        y1 = f"=IF(C42>0,C46/C42,0)"
        y2 = f"=IF(D42>0,D46/D42,0)"
        tot = f"=IF(E42>0,E46/E42,0)"
    elif label == "RETURNS & REPLACEMENTS" or label == "ADJUSTED P&L (After Returns)":
        # Section headers — no formulas
        continue
    else:
        continue

    for ci, fml in [(3, y1), (4, y2), (5, tot)]:
        cell = ws3.cell(row=row, column=ci)
        cell.value = fml; cell.font = f; cell.number_format = fmt; cell.border = thin_border
        if fill:
            cell.fill = fill


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: SCENARIO PLAYBOOK
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Scenario Notes")
ws4.sheet_properties.tabColor = "7B1FA2"
ws4.column_dimensions["A"].width = 3
ws4.column_dimensions["B"].width = 50
ws4.column_dimensions["C"].width = 25
ws4.column_dimensions["D"].width = 25
ws4.column_dimensions["E"].width = 25

sc(ws4, 1, 2, "Scenario Playbook", font=title_font)
sc(ws4, 2, 2, "Copy the workbook, change Assumptions, compare results", font=Font(name="Calibri", italic=True, size=10, color="666666"))

R = 4
section_header(ws4, R, 2, 5, ["Assumption", "Conservative", "Base Case", "Optimistic"])

scenarios = [
    ("10-Year Charm Price", "$24.99", "$29.99", "$34.99"),
    ("15-Year Charm Price", "$39.99", "$44.99", "$54.99"),
    ("Perpetual Charm Price", "$59.99", "$69.99", "$89.99"),
    ("Starting 10-Year Units/Mo", "50", "100", "200"),
    ("Starting 15-Year Units/Mo", "20", "40", "80"),
    ("Starting Perpetual Units/Mo", "5", "15", "30"),
    ("MoM Growth (10-Year)", "5%", "8%", "12%"),
    ("MoM Growth (15-Year)", "6%", "10%", "15%"),
    ("MoM Growth (Perpetual)", "8%", "12%", "18%"),
    ("COGS per Charm", "$10.00", "$8.50", "$7.00"),
    ("Avg Content Size (MB)", "20", "35.9", "60"),
    ("Views/Charm/Mo (New Period)", "4", "8", "15"),
    ("Views/Charm/Mo (Long-Tail)", "1", "2", "5"),
    ("Video Mix %", "40%", "55%", "70%"),
    ("Marketing Budget (M1)", "$250", "$500", "$1,000"),
    ("Extend Memory Attach Rate", "1%", "2%", "4%"),
    ("Gift Wrap Attach Rate", "8%", "15%", "25%"),
]

for i, (label, cons, base, opt) in enumerate(scenarios):
    row = R + 1 + i
    sc(ws4, row, 2, label, font=normal_font, border=thin_border)
    sc(ws4, row, 3, cons, font=normal_font, border=thin_border, alignment=Alignment(horizontal="center"))
    sc(ws4, row, 4, base, font=bold_font, border=thin_border, alignment=Alignment(horizontal="center"), fill=green_fill)
    sc(ws4, row, 5, opt, font=normal_font, border=thin_border, alignment=Alignment(horizontal="center"))

r2 = R + len(scenarios) + 2
sc(ws4, r2, 2, "HOW TO USE:", font=section_font)
sc(ws4, r2+1, 2, "1. Base Case values match the Assumptions sheet defaults", font=normal_font)
sc(ws4, r2+2, 2, "2. Go to Assumptions sheet and change the yellow cells", font=normal_font)
sc(ws4, r2+3, 2, "3. Projections and Annual Summary auto-update", font=normal_font)
sc(ws4, r2+4, 2, "4. File > Save As to create copies for each scenario", font=normal_font)
sc(ws4, r2+5, 2, "5. Key infrastructure levers: content size, views/charm, video mix %", font=normal_font)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: PER-CHARM UNIT ECONOMICS (No Free Tier)
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Per-Charm Costs")
ws5.sheet_properties.tabColor = "D32F2F"
ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 55
ws5.column_dimensions["C"].width = 22
ws5.column_dimensions["D"].width = 22
ws5.column_dimensions["E"].width = 22
ws5.column_dimensions["F"].width = 40

A = "Assumptions"
cost_6dp = '$#,##0.000000'
cost_4dp = '$#,##0.0000'

sc(ws5, 1, 2, "Per-Charm Hosting Cost Breakdown (No Free Tier)", font=title_font)
sc(ws5, 2, 2, "What it costs to host ONE charm — raw service rates, no free tier credits applied", font=Font(name="Calibri", italic=True, size=10, color="666666"))
sc(ws5, 3, 2, "All values linked to Assumptions sheet — change inputs there to update these costs", font=Font(name="Calibri", italic=True, size=10, color="666666"))

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Content & Activity Profile for 1 Charm
# ────────────────────────────────────────────────────────────────────────────
R = 5
sc(ws5, R, 2, "SINGLE CHARM PROFILE", font=section_font)
section_header(ws5, R+1, 2, 4, ["Metric", "Value", "Unit"])

R = 7
sc(ws5, R, 2, "Content size (weighted avg)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"={A}!C33", normal_font, num_2dp)
note(ws5, R, 4, "MB")
note(ws5, R, 6, f"=TEXT({A}!D29*100,\"0\")&\"% video (\"&{A}!C29&\" MB), \"&TEXT({A}!D30*100,\"0\")&\"% image (\"&{A}!C30&\" MB), \"&TEXT({A}!D31*100,\"0\")&\"% audio (\"&{A}!C31&\" MB)\"")
# simpler note
sc(ws5, R, 6, None)
note(ws5, R, 6, "From Assumptions: video/image/audio weighted mix")

R = 8
sc(ws5, R, 2, "Content size in GB", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=C7/1024", normal_font, '#,##0.000000')
note(ws5, R, 4, "GB")

R = 9
sc(ws5, R, 2, "Views per month (first 3 months — novelty)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"={A}!C38", normal_font, num_fmt)
note(ws5, R, 4, "views/month")

R = 10
sc(ws5, R, 2, "Views per month (after 3 months — long-tail)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"={A}!C39", normal_font, num_fmt)
note(ws5, R, 4, "views/month")

R = 11
sc(ws5, R, 2, "Blended avg views/month (assuming mostly long-tail)", font=bold_font, border=thin_border)
# Weighted: 3 months of novelty + 9 months of long-tail = annual average / 12
formula(ws5, R, 3, f"=ROUND((C9*3+C10*9)/12,1)", bold_font, num_1dp)
note(ws5, R, 4, "views/month")
note(ws5, R, 6, "Year-weighted: 3 months novelty + 9 months long-tail")

R = 12
sc(ws5, R, 2, "% of views requiring glyph verification", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"={A}!C40", normal_font, pct_fmt)

R = 13
sc(ws5, R, 2, "API calls per view (GetCharm + glyph if needed)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=1+C12", normal_font, num_2dp)
note(ws5, R, 6, "1 GetCharm call per view + glyph verify probability")

R = 14
sc(ws5, R, 2, "Table reads per view", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"={A}!C43", normal_font, num_fmt)

R = 15
sc(ws5, R, 2, "Table writes per view (glyph attempts + rate-limit log)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=ROUND(C12*2+0.1,1)", normal_font, num_1dp)
note(ws5, R, 6, "2 writes per glyph verify + 10% sampled request logging")

# ────────────────────────────────────────────────────────────────────────────
# SECTION: One-Time Setup Costs (Claim → Configure → Upload → Finalize)
# ────────────────────────────────────────────────────────────────────────────
R = 17
sc(ws5, R, 2, "ONE-TIME SETUP COSTS (per charm — claim through finalize)", font=section_font)
section_header(ws5, R+1, 2, 5, ["Service / Operation", "Cost", "Calculation", "Notes"])

# Azure Functions — lifecycle API calls
R = 19
sc(ws5, R, 2, "Azure Functions: Lifecycle API Calls", font=normal_font, border=thin_border)
# cost = (lifecycle_calls / 1M) * rate + (lifecycle_calls * duration_s * mem_GB) * rate
formula(ws5, R, 3,
    f"=({A}!C41/1000000)*{A}!C69 + {A}!C41*({A}!C71/1000)*({A}!C72/1024)*{A}!C70",
    normal_font, cost_6dp)
note(ws5, R, 4, "5 calls x exec + compute cost")
note(ws5, R, 6, "claim + configure + get-upload-urls + finalize + preview")

# R2 Class A — initial upload writes
R = 20
sc(ws5, R, 2, "Cloudflare R2: Upload Writes (Class A)", font=normal_font, border=thin_border)
# 2 write ops * (1 + re-upload rate) / 1M * rate
formula(ws5, R, 3,
    f"=(2*(1+{A}!C35)/1000000)*{A}!C56",
    normal_font, cost_6dp)
note(ws5, R, 4, "2 PUT ops x (1+re-upload rate)")
note(ws5, R, 6, "PUT blob + list verify; re-uploads during 14-day settling")

# Azure Blob — backup upload writes
R = 21
sc(ws5, R, 2, "Azure Blob: Backup Upload Writes", font=normal_font, border=thin_border)
formula(ws5, R, 3,
    f"=(2*(1+{A}!C35)/10000)*{A}!C48",
    normal_font, cost_6dp)
note(ws5, R, 4, "2 ops x (1+re-upload rate) / 10K")

# Azure Table — setup entities
R = 22
sc(ws5, R, 2, "Azure Table: Setup Entity Writes", font=normal_font, border=thin_border)
formula(ws5, R, 3,
    f"=({A}!C42/10000)*{A}!C64",
    normal_font, cost_6dp)
note(ws5, R, 4, "8 writes / 10K x txn rate")
note(ws5, R, 6, "charm + user-charm + profile + request log entities")

# Azure Table — entity storage (one-time provisioning, ongoing is below)
R = 23
sc(ws5, R, 2, "Azure Table: Entity Storage (first month)", font=normal_font, border=thin_border)
formula(ws5, R, 3,
    f"=({A}!C65/1024/1024)*{A}!C63",
    normal_font, cost_6dp)
note(ws5, R, 4, "~2 KB entity / GB x rate")

# TOTAL ONE-TIME
R = 24
sc(ws5, R, 2, "TOTAL ONE-TIME SETUP COST", font=bold_font, border=thin_border, fill=red_fill)
formula(ws5, R, 3, "=SUM(C19:C23)", bold_font, cost_4dp)
style_range(ws5, R, 2, 5, fill=red_fill)

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Monthly Ongoing Costs (Storage + Playback Activity)
# ────────────────────────────────────────────────────────────────────────────
R = 26
sc(ws5, R, 2, "MONTHLY ONGOING COSTS (per charm — storage + playback)", font=section_font)
note(ws5, R, 6, "Assumes blended avg views/month from row 11")
section_header(ws5, R+1, 2, 5, ["Service / Operation", "$/Month", "Calculation", "Notes"])

# --- CLOUDFLARE R2 ---
R = 28
sc(ws5, R, 2, "CLOUDFLARE R2", font=Font(name="Calibri", bold=True, size=11, color=MED_BLUE), border=thin_border)
style_range(ws5, R, 2, 5, fill=PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"))

R = 29
sc(ws5, R, 2, "  R2: Storage (content GB x rate)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=C8*{A}!C55", normal_font, cost_6dp)
note(ws5, R, 4, "GB x $0.015/GB/mo")

R = 30
sc(ws5, R, 2, "  R2: Class B Reads — playback downloads", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=(C11/1000000)*{A}!C57", normal_font, cost_6dp)
note(ws5, R, 4, "views/mo / 1M x rate")
note(ws5, R, 6, "Every view = 1 GET from R2")

R = 31
sc(ws5, R, 2, "  R2: Egress / Bandwidth", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=C11*C8*{A}!C58", normal_font, cost_6dp)
note(ws5, R, 4, "$0.00 — zero egress")
note(ws5, R, 6, "Cloudflare's key advantage: zero bandwidth cost")

R = 32
sc(ws5, R, 2, "  R2 SUBTOTAL", font=bold_font, border=thin_border)
formula(ws5, R, 3, "=C29+C30+C31", bold_font, cost_6dp)

# --- AZURE BLOB COOL ---
R = 34
sc(ws5, R, 2, "AZURE BLOB STORAGE — COOL TIER", font=Font(name="Calibri", bold=True, size=11, color=MED_BLUE), border=thin_border)
style_range(ws5, R, 2, 5, fill=PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"))

R = 35
sc(ws5, R, 2, "  Azure Blob: Storage (backup mirror — same GB)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=C8*{A}!C47", normal_font, cost_6dp)
note(ws5, R, 4, "GB x $0.01/GB/mo")

R = 36
sc(ws5, R, 2, "  Azure Blob: Read Ops (fallback — 2% of views)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=(C11*0.02/10000)*{A}!C49", normal_font, cost_6dp)
note(ws5, R, 4, "2% fallback reads")
note(ws5, R, 6, "Only when R2 is unavailable; client falls back to Azure")

R = 37
sc(ws5, R, 2, "  Azure Blob: Data Retrieval (cool tier penalty)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=C11*0.02*C8*{A}!C50", normal_font, cost_6dp)
note(ws5, R, 4, "2% fallback x GB x retrieval rate")
note(ws5, R, 6, "Cool tier charges per-GB on read; hot tier does not")

R = 38
sc(ws5, R, 2, "  Azure Blob: Egress (fallback downloads)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=C11*0.02*C8*{A}!C51", normal_font, cost_6dp)
note(ws5, R, 4, "2% fallback x GB x egress rate")
note(ws5, R, 6, "$0.087/GB — only on Azure-direct downloads")

R = 39
sc(ws5, R, 2, "  Azure Blob SUBTOTAL", font=bold_font, border=thin_border)
formula(ws5, R, 3, "=C35+C36+C37+C38", bold_font, cost_6dp)

# --- AZURE TABLE STORAGE ---
R = 41
sc(ws5, R, 2, "AZURE TABLE STORAGE", font=Font(name="Calibri", bold=True, size=11, color=MED_BLUE), border=thin_border)
style_range(ws5, R, 2, 5, fill=PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"))

R = 42
sc(ws5, R, 2, "  Table: Entity Storage (ongoing)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=({A}!C65/1024/1024)*{A}!C63", normal_font, cost_6dp)
note(ws5, R, 4, "~2 KB entity set / month")

R = 43
sc(ws5, R, 2, "  Table: Read Transactions (views x reads/view)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=(C11*C14/10000)*{A}!C64", normal_font, cost_6dp)
note(ws5, R, 4, "views x 3 reads / 10K x rate")

R = 44
sc(ws5, R, 2, "  Table: Write Transactions (glyph + logging)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=(C11*C15/10000)*{A}!C64", normal_font, cost_6dp)
note(ws5, R, 4, "views x writes/view / 10K x rate")

R = 45
sc(ws5, R, 2, "  Azure Table SUBTOTAL", font=bold_font, border=thin_border)
formula(ws5, R, 3, "=C42+C43+C44", bold_font, cost_6dp)

# --- AZURE FUNCTIONS ---
R = 47
sc(ws5, R, 2, "AZURE FUNCTIONS (Compute)", font=Font(name="Calibri", bold=True, size=11, color=MED_BLUE), border=thin_border)
style_range(ws5, R, 2, 5, fill=PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"))

R = 48
sc(ws5, R, 2, "  Functions: Execution Cost (view + glyph calls)", font=normal_font, border=thin_border)
# Total API calls per charm per month = blended views * calls_per_view
formula(ws5, R, 3, f"=(C11*C13/1000000)*{A}!C69", normal_font, cost_6dp)
note(ws5, R, 4, "calls / 1M x $0.20")

R = 49
sc(ws5, R, 2, "  Functions: Compute GB-seconds", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=C11*C13*({A}!C71/1000)*({A}!C72/1024)*{A}!C70", normal_font, cost_6dp)
note(ws5, R, 4, "calls x duration x memory x rate")
note(ws5, R, 6, "~200ms avg x 128MB per invocation")

R = 50
sc(ws5, R, 2, "  Azure Functions SUBTOTAL", font=bold_font, border=thin_border)
formula(ws5, R, 3, "=C48+C49", bold_font, cost_6dp)

# --- ENTRA CIAM ---
R = 52
sc(ws5, R, 2, "ENTRA CIAM (Authentication)", font=Font(name="Calibri", bold=True, size=11, color=MED_BLUE), border=thin_border)
style_range(ws5, R, 2, 5, fill=PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"))

R = 53
sc(ws5, R, 2, "  CIAM: Per-MAU cost (1 charm ≈ 0.3 MAU avg)", font=normal_font, border=thin_border)
formula(ws5, R, 3, f"=0.3*{A}!C79", normal_font, cost_6dp)
note(ws5, R, 4, "0.3 MAU x $0.0025/MAU")
note(ws5, R, 6, "Not every charm = unique monthly active user")

R = 54
sc(ws5, R, 2, "  CIAM SUBTOTAL", font=bold_font, border=thin_border)
formula(ws5, R, 3, "=C53", bold_font, cost_6dp)

# ────────────────────────────────────────────────────────────────────────────
# TOTAL MONTHLY ONGOING
# ────────────────────────────────────────────────────────────────────────────
R = 56
sc(ws5, R, 2, "TOTAL MONTHLY HOSTING COST (per charm)", font=Font(name="Calibri", bold=True, size=13, color=DARK_BLUE), border=thin_border, fill=red_fill)
formula(ws5, R, 3, "=C32+C39+C45+C50+C54", Font(name="Calibri", bold=True, size=13, color=DARK_BLUE), cost_4dp)
ws5.cell(row=R, column=3).fill = red_fill
style_range(ws5, R, 2, 5, fill=red_fill)

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Annual & Lifetime Rollups
# ────────────────────────────────────────────────────────────────────────────
R = 58
sc(ws5, R, 2, "ANNUAL & LIFETIME COST SUMMARY (per charm)", font=section_font)
section_header(ws5, R+1, 2, 5, ["Time Horizon", "Hosting Cost", "Incl. Setup", "Notes"])

R = 60
sc(ws5, R, 2, "Annual Hosting Cost (monthly x 12)", font=bold_font, border=thin_border)
formula(ws5, R, 3, "=C56*12", bold_font, cost_4dp)
formula(ws5, R, 4, "=C56*12+C24", bold_font, cost_4dp)
note(ws5, R, 6, "Pure recurring hosting — no COGS, no shipping")

R = 61
sc(ws5, R, 2, "Year 1 Total (setup + 12 months hosting)", font=bold_font, border=thin_border)
formula(ws5, R, 3, "=C56*12", bold_font, cost_4dp)
formula(ws5, R, 4, "=C24+C56*12", bold_font, cost_4dp)
note(ws5, R, 6, "First year is highest — includes one-time setup")

R = 63
sc(ws5, R, 2, "LIFETIME HOSTING BY TIER", font=section_font)
section_header(ws5, R+1, 2, 6, ["Tier", "Hosting Only", "Incl. Setup", "As % of Retail Price", "Hosting Profit/Loss"])

R = 65
sc(ws5, R, 2, "10-Year Charm (120 months hosting)", font=normal_font, border=thin_border)
formula(ws5, R, 3, "=C56*120", normal_font, currency_fmt)
formula(ws5, R, 4, "=C24+C56*120", normal_font, currency_fmt)
formula(ws5, R, 5, f"=IF({A}!C6>0,D65/{A}!C6,0)", normal_font, pct_fmt)
formula(ws5, R, 6, f"={A}!C6-D65", normal_font, currency_fmt)
note(ws5, R, 6, None)
sc(ws5, R, 6, None)
# Use a formula for this
cell = ws5.cell(row=R, column=6)
cell.value = f"={A}!C6-D65"
cell.font = bold_font
cell.number_format = currency_fmt
cell.border = thin_border

R = 66
sc(ws5, R, 2, "15-Year Charm (180 months hosting)", font=normal_font, border=thin_border)
formula(ws5, R, 3, "=C56*180", normal_font, currency_fmt)
formula(ws5, R, 4, "=C24+C56*180", normal_font, currency_fmt)
formula(ws5, R, 5, f"=IF({A}!C7>0,D66/{A}!C7,0)", normal_font, pct_fmt)
cell = ws5.cell(row=R, column=6)
cell.value = f"={A}!C7-D66"
cell.font = bold_font; cell.number_format = currency_fmt; cell.border = thin_border

R = 67
sc(ws5, R, 2, "Retail / Perpetual Charm (30-year est = 360 months)", font=normal_font, border=thin_border)
formula(ws5, R, 3, "=C56*360", normal_font, currency_fmt)
formula(ws5, R, 4, "=C24+C56*360", normal_font, currency_fmt)
formula(ws5, R, 5, f"=IF({A}!C8>0,D67/{A}!C8,0)", normal_font, pct_fmt)
cell = ws5.cell(row=R, column=6)
cell.value = f"={A}!C8-D67"
cell.font = bold_font; cell.number_format = currency_fmt; cell.border = thin_border

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Service-by-Service Monthly Percentage Breakdown
# ────────────────────────────────────────────────────────────────────────────
R = 69
sc(ws5, R, 2, "MONTHLY COST — SERVICE BREAKDOWN (%)", font=section_font)
section_header(ws5, R+1, 2, 5, ["Service", "$/Month", "% of Total", "Key Driver"])

R = 71
sc(ws5, R, 2, "Cloudflare R2", font=normal_font, border=thin_border)
formula(ws5, R, 3, "=C32", normal_font, cost_6dp)
formula(ws5, R, 4, "=IF(C56>0,C32/C56,0)", normal_font, pct_fmt)
note(ws5, R, 5, "Storage volume (GB per charm)")

R = 72
sc(ws5, R, 2, "Azure Blob (Cool Backup)", font=normal_font, border=thin_border)
formula(ws5, R, 3, "=C39", normal_font, cost_6dp)
formula(ws5, R, 4, "=IF(C56>0,C39/C56,0)", normal_font, pct_fmt)
note(ws5, R, 5, "Storage + fallback reads")

R = 73
sc(ws5, R, 2, "Azure Table Storage", font=normal_font, border=thin_border)
formula(ws5, R, 3, "=C45", normal_font, cost_6dp)
formula(ws5, R, 4, "=IF(C56>0,C45/C56,0)", normal_font, pct_fmt)
note(ws5, R, 5, "Transaction volume (views x ops)")

R = 74
sc(ws5, R, 2, "Azure Functions", font=normal_font, border=thin_border)
formula(ws5, R, 3, "=C50", normal_font, cost_6dp)
formula(ws5, R, 4, "=IF(C56>0,C50/C56,0)", normal_font, pct_fmt)
note(ws5, R, 5, "API call volume")

R = 75
sc(ws5, R, 2, "Entra CIAM", font=normal_font, border=thin_border)
formula(ws5, R, 3, "=C54", normal_font, cost_6dp)
formula(ws5, R, 4, "=IF(C56>0,C54/C56,0)", normal_font, pct_fmt)
note(ws5, R, 5, "Monthly active users")

R = 76
sc(ws5, R, 2, "TOTAL", font=bold_font, border=thin_border, fill=red_fill)
formula(ws5, R, 3, "=C56", bold_font, cost_4dp)
ws5.cell(row=R, column=3).fill = red_fill
formula(ws5, R, 4, "=SUM(D71:D75)", bold_font, pct_fmt)
ws5.cell(row=R, column=4).fill = red_fill

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Content Size Sensitivity (what if charm is bigger/smaller?)
# ────────────────────────────────────────────────────────────────────────────
R = 78
sc(ws5, R, 2, "CONTENT SIZE SENSITIVITY — Monthly Cost at Different Sizes", font=section_font)
note(ws5, R, 6, "Holding view rate constant, varying content size")
section_header(ws5, R+1, 2, 6, ["Content Size", "R2 Storage", "Azure Blob", "Functions+Table", "Total Monthly", "Annual"])

sizes_mb = [6, 15, 35.85, 55, 100, 150]
for i, size in enumerate(sizes_mb):
    row = 80 + i
    label = (f"{size} MB" + (
        " (audio only)" if size == 6
        else " (image gallery)" if size == 15
        else " (weighted avg — base case)" if size == 35.85
        else " (avg video — iPhone 1080p mix)" if size == 55
        else " (4K / H.264 heavy)" if size == 100
        else " (max — 150 MB limit)" ))
    sc(ws5, row, 2, label, font=bold_font if size == 35.85 else normal_font, border=thin_border)
    # R2 storage = size/1024 * R2 rate
    formula(ws5, row, 3, f"=({size}/1024)*{A}!C55", normal_font, cost_6dp)
    # Azure blob = size/1024 * Azure rate + fallback read/egress
    formula(ws5, row, 4,
        f"=({size}/1024)*{A}!C47 + (C11*0.02/10000)*{A}!C49 + C11*0.02*({size}/1024)*{A}!C50 + C11*0.02*({size}/1024)*{A}!C51",
        normal_font, cost_6dp)
    # Functions + Table (same regardless of content size, driven by views)
    formula(ws5, row, 5, "=C50+C45", normal_font, cost_6dp)
    # Total monthly
    cell = ws5.cell(row=row, column=6)
    cell.value = f"=C{row}+D{row}+E{row}+C54"
    cell.font = bold_font if size == 35.85 else normal_font
    cell.number_format = cost_4dp
    cell.border = thin_border

    if size == 35.85:
        style_range(ws5, row, 2, 6, fill=green_fill)

# Total annual column
R = 78
sc(ws5, R+1, 7, "Annual (x12)", font=header_font, fill=header_fill, alignment=Alignment(horizontal="center"))
ws5.column_dimensions["G"].width = 16
for i in range(len(sizes_mb)):
    row = 80 + i
    cell = ws5.cell(row=row, column=7)
    cell.value = f"=F{row}*12"
    cell.font = normal_font
    cell.number_format = cost_4dp
    cell.border = thin_border


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6: REVENUE RECOGNITION — Cash vs Hybrid vs Straight-Line
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Revenue Recognition")
ws6.sheet_properties.tabColor = "1565C0"
ws6.column_dimensions["A"].width = 3
ws6.column_dimensions["B"].width = 52

sc(ws6, 1, 2, "Revenue Recognition — Cash vs Hybrid vs Straight-Line", font=title_font)
sc(ws6, 2, 2, "Compares when revenue is recognized under each method + deferred revenue liability", font=Font(name="Calibri", italic=True, size=10, color="666666"))

# Month headers row 4
R = 4
sc(ws6, R, 2, "", font=header_font, fill=header_fill)
for m in range(1, 25):
    col = m + 2
    ws6.column_dimensions[get_column_letter(col)].width = 14
    sc(ws6, R, col, f"Month {m}", font=header_font, fill=header_fill, alignment=Alignment(horizontal="center"))

PRJ = "'24-Month Projections'"
PCC = "'Per-Charm Costs'"
ASM2 = "Assumptions"
orange_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
blue_light = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Per-Charm Revenue Split (reference from Per-Charm Costs & Assumptions)
# ────────────────────────────────────────────────────────────────────────────
R = 6
sc(ws6, R, 2, "PER-CHARM REVENUE SPLIT — HYBRID METHOD", font=section_font)
sc(ws6, R, 3, "10-Year", font=header_font, fill=header_fill, alignment=Alignment(horizontal="center"))
sc(ws6, R, 4, "15-Year", font=header_font, fill=header_fill, alignment=Alignment(horizontal="center"))
sc(ws6, R, 5, "Perpetual", font=header_font, fill=header_fill, alignment=Alignment(horizontal="center"))
note(ws6, R, 6, "Lifetime months:")
sc(ws6, R, 7, 120, font=bold_font, border=thin_border)
sc(ws6, R, 8, 180, font=bold_font, border=thin_border)
sc(ws6, R, 9, 360, font=bold_font, border=thin_border)

# Row 7: Sale price
R = 7
sc(ws6, R, 2, "Sale Price", font=normal_font, border=thin_border)
formula(ws6, R, 3, f"={ASM2}!C6", normal_font, currency_fmt)
formula(ws6, R, 4, f"={ASM2}!C7", normal_font, currency_fmt)
formula(ws6, R, 5, f"={ASM2}!C8", normal_font, currency_fmt)

# Row 8: Upfront costs (recognized at sale) = COGS + shipping + processing% * price + one-time hosting
R = 8
sc(ws6, R, 2, "Upfront Costs (COGS + ship + processing + setup)", font=normal_font, border=thin_border)
formula(ws6, R, 3, f"={ASM2}!D6+{ASM2}!C86+C7*{ASM2}!C87+{PCC}!C24", normal_font, currency_fmt)
formula(ws6, R, 4, f"={ASM2}!D7+{ASM2}!C86+D7*{ASM2}!C87+{PCC}!C24", normal_font, currency_fmt)
formula(ws6, R, 5, f"={ASM2}!D8+{ASM2}!C86+E7*{ASM2}!C87+{PCC}!C24", normal_font, currency_fmt)

# Row 9: Upfront revenue recognized (= upfront costs; match cost at sale)
R = 9
sc(ws6, R, 2, "Hybrid: Revenue Recognized at Sale", font=bold_font, border=thin_border)
formula(ws6, R, 3, "=C8", bold_font, currency_fmt)
formula(ws6, R, 4, "=D8", bold_font, currency_fmt)
formula(ws6, R, 5, "=E8", bold_font, currency_fmt)
note(ws6, R, 6, "Covers COGS + fulfillment + payment fee + setup hosting")

# Row 10: Deferred portion (price - upfront)
R = 10
sc(ws6, R, 2, "Hybrid: Deferred Revenue (to escrow)", font=bold_font, border=thin_border, fill=orange_fill)
formula(ws6, R, 3, "=C7-C9", bold_font, currency_fmt)
formula(ws6, R, 4, "=D7-D9", bold_font, currency_fmt)
formula(ws6, R, 5, "=E7-E9", bold_font, currency_fmt)
style_range(ws6, R, 2, 5, fill=orange_fill)
note(ws6, R, 6, "Goes to escrow / deferred revenue liability on balance sheet")

# Row 11: Monthly recognition from deferred pool
R = 11
sc(ws6, R, 2, "Hybrid: Monthly Recognition from Deferred", font=bold_font, border=thin_border)
formula(ws6, R, 3, "=C10/G6", bold_font, currency_micro)  # deferred / 120 months
formula(ws6, R, 4, "=D10/H6", bold_font, currency_micro)
formula(ws6, R, 5, "=E10/I6", bold_font, currency_micro)
note(ws6, R, 6, "Deferred portion ÷ lifetime months")

# Row 12: Straight-line monthly recognition (full price / lifetime)
R = 12
sc(ws6, R, 2, "Straight-Line: Monthly Recognition (all deferred)", font=bold_font, border=thin_border, fill=blue_light)
formula(ws6, R, 3, "=C7/G6", bold_font, currency_micro)
formula(ws6, R, 4, "=D7/H6", bold_font, currency_micro)
formula(ws6, R, 5, "=E7/I6", bold_font, currency_micro)
style_range(ws6, R, 2, 5, fill=blue_light)
note(ws6, R, 6, "Full price ÷ lifetime months — most conservative")

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Cash Collected (monthly) — same as current model
# ────────────────────────────────────────────────────────────────────────────
R = 14
sc(ws6, R, 2, "CASH COLLECTED (Point of Sale)", font=section_font)

# Need cumulative by tier for recognition formulas — track these inline
# Row 15: 10-Year cumulative
R = 15; sc(ws6, R, 2, "10-Year Charms — Cumulative Sold", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    if m == 1:
        f = f"={PRJ}!{c}6"
    else:
        prev = get_column_letter(col - 1)
        f = f"={prev}{R}+{PRJ}!{c}6"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = num_fmt; cell.border = thin_border

# Row 16: 15-Year cumulative
R = 16; sc(ws6, R, 2, "15-Year Charms — Cumulative Sold", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    if m == 1:
        f = f"={PRJ}!{c}7"
    else:
        prev = get_column_letter(col - 1)
        f = f"={prev}{R}+{PRJ}!{c}7"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = num_fmt; cell.border = thin_border

# Row 17: Perpetual cumulative
R = 17; sc(ws6, R, 2, "Perpetual Charms — Cumulative Sold", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    if m == 1:
        f = f"={PRJ}!{c}8"
    else:
        prev = get_column_letter(col - 1)
        f = f"={prev}{R}+{PRJ}!{c}8"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = num_fmt; cell.border = thin_border

# Row 18: Total Cash Collected (charm sales only — no upsells for simplicity)
R = 18; sc(ws6, R, 2, "Charm Cash Collected (month)", font=bold_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={PRJ}!{c}6*$C$7+{PRJ}!{c}7*$D$7+{PRJ}!{c}8*$E$7"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 19: Upsell Cash (recognized immediately — Gift Wrap + Upgrade)
R = 19; sc(ws6, R, 2, "Upsell Cash Collected (recognized immediately)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={PRJ}!{c}49+{PRJ}!{c}50+{PRJ}!{c}51"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 20: TOTAL CASH
R = 20; sc(ws6, R, 2, "TOTAL CASH COLLECTED", font=bold_font, border=thin_border, fill=green_fill)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={c}18+{c}19"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = green_fill
ws6.cell(row=R, column=2).fill = green_fill

# ────────────────────────────────────────────────────────────────────────────
# SECTION: HYBRID Recognized Revenue
# ────────────────────────────────────────────────────────────────────────────
R = 22; sc(ws6, R, 2, "HYBRID METHOD — Recognized Revenue", font=section_font)
note(ws6, R, 6, None)
sc(ws6, R, 6, "COGS portion at sale + deferred over lifetime", font=Font(name="Calibri", italic=True, size=10, color="666666"))

# Row 23: Upfront portion recognized this month (new sales * upfront per tier)
R = 23; sc(ws6, R, 2, "  Upfront Recognition (new sales x upfront portion)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={PRJ}!{c}6*$C$9+{PRJ}!{c}7*$D$9+{PRJ}!{c}8*$E$9"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 24: Deferred recognition this month (all cumulative charms * monthly recognition rate)
# Each charm sold in any prior month contributes its monthly rate.
# Within 24 months, no charms expire (shortest = 120 months), so all cumulative contribute.
R = 24; sc(ws6, R, 2, "  Deferred Recognition (cumulative charms x monthly rate)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={c}15*$C$11+{c}16*$D$11+{c}17*$E$11"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 25: Upsell (immediate)
R = 25; sc(ws6, R, 2, "  Upsell Revenue (immediate recognition)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={c}19"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 26: TOTAL HYBRID RECOGNIZED
R = 26; sc(ws6, R, 2, "HYBRID RECOGNIZED REVENUE", font=bold_font, border=thin_border, fill=orange_fill)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={c}23+{c}24+{c}25"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = orange_fill
ws6.cell(row=R, column=2).fill = orange_fill

# ────────────────────────────────────────────────────────────────────────────
# SECTION: STRAIGHT-LINE Recognized Revenue
# ────────────────────────────────────────────────────────────────────────────
R = 28; sc(ws6, R, 2, "STRAIGHT-LINE METHOD — Recognized Revenue", font=section_font)
sc(ws6, R, 6, "Full price spread evenly over charm lifetime", font=Font(name="Calibri", italic=True, size=10, color="666666"))

# Row 29: Monthly SL recognition (all cumulative charms * SL monthly rate per tier)
R = 29; sc(ws6, R, 2, "  Charm Revenue (cumulative x monthly rate)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={c}15*$C$12+{c}16*$D$12+{c}17*$E$12"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 30: Upsell (immediate, same as hybrid)
R = 30; sc(ws6, R, 2, "  Upsell Revenue (immediate recognition)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={c}19"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 31: TOTAL STRAIGHT-LINE RECOGNIZED
R = 31; sc(ws6, R, 2, "STRAIGHT-LINE RECOGNIZED REVENUE", font=bold_font, border=thin_border, fill=blue_light)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={c}29+{c}30"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = blue_light
ws6.cell(row=R, column=2).fill = blue_light

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Side-by-Side Comparison
# ────────────────────────────────────────────────────────────────────────────
R = 33; sc(ws6, R, 2, "SIDE-BY-SIDE COMPARISON", font=section_font)

# Row 34: Cash
R = 34; sc(ws6, R, 2, "Cash Collected", font=bold_font, border=thin_border, fill=green_fill)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col); cell.value = f"={c}20"
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = green_fill
ws6.cell(row=R, column=2).fill = green_fill

# Row 35: Hybrid
R = 35; sc(ws6, R, 2, "Hybrid Recognized", font=bold_font, border=thin_border, fill=orange_fill)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col); cell.value = f"={c}26"
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = orange_fill
ws6.cell(row=R, column=2).fill = orange_fill

# Row 36: Straight-line
R = 36; sc(ws6, R, 2, "Straight-Line Recognized", font=bold_font, border=thin_border, fill=blue_light)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col); cell.value = f"={c}31"
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = blue_light
ws6.cell(row=R, column=2).fill = blue_light

# Row 37: Cash - Hybrid gap
R = 37; sc(ws6, R, 2, "Cash vs Hybrid Gap (unrecognized cash)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col); cell.value = f"={c}34-{c}35"
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 38: Cash - SL gap
R = 38; sc(ws6, R, 2, "Cash vs Straight-Line Gap (unrecognized cash)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col); cell.value = f"={c}34-{c}36"
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Deferred Revenue Liability (Balance Sheet)
# ────────────────────────────────────────────────────────────────────────────
R = 40; sc(ws6, R, 2, "DEFERRED REVENUE LIABILITY (Balance Sheet)", font=section_font)
note(ws6, R, 6, "This is your escrow — cash received but not yet earned")

# --- HYBRID ---
R = 42; sc(ws6, R, 2, "HYBRID — Deferred Revenue Balance", font=bold_font, border=thin_border, fill=orange_fill)
note(ws6, R, 6, "Opening + new deferrals - monthly recognition = closing")
style_range(ws6, R, 2, 26, fill=orange_fill)

# Row 43: New deferrals this month
R = 43; sc(ws6, R, 2, "  + New Deferrals (charm sales x deferred portion)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={PRJ}!{c}6*$C$10+{PRJ}!{c}7*$D$10+{PRJ}!{c}8*$E$10"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 44: Recognition (drawdown)
R = 44; sc(ws6, R, 2, "  - Monthly Recognition (from deferred pool)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col); cell.value = f"={c}24"
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 45: Return reversals (refunded charms — remove their deferred portion)
R = 45; sc(ws6, R, 2, "  - Return Reversals (refunded charms x deferred portion)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    # Returned units * weighted avg deferred per charm
    # Approximate: returned units * (total new deferrals this month / total sales this month)
    f = f"=IF({PRJ}!{c}9>0, {PRJ}!{c}88*({c}43/{PRJ}!{c}9), 0)"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 46: Closing balance
R = 46; sc(ws6, R, 2, "  = HYBRID Deferred Revenue Balance", font=bold_font, border=thin_border, fill=orange_fill)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    if m == 1:
        f = f"={c}43-{c}44-{c}45"
    else:
        prev = get_column_letter(col - 1)
        f = f"={prev}{R}+{c}43-{c}44-{c}45"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = orange_fill
ws6.cell(row=R, column=2).fill = orange_fill

# --- STRAIGHT-LINE ---
R = 48; sc(ws6, R, 2, "STRAIGHT-LINE — Deferred Revenue Balance", font=bold_font, border=thin_border, fill=blue_light)
style_range(ws6, R, 2, 26, fill=blue_light)

# Row 49: New deferrals (entire sale price)
R = 49; sc(ws6, R, 2, "  + New Deferrals (full charm price)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"={PRJ}!{c}6*$C$7+{PRJ}!{c}7*$D$7+{PRJ}!{c}8*$E$7"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 50: Recognition (drawdown)
R = 50; sc(ws6, R, 2, "  - Monthly Recognition", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col); cell.value = f"={c}29"
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 51: Return reversals (full charm price reversed)
R = 51; sc(ws6, R, 2, "  - Return Reversals (refunded charms x full price)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    f = f"=IF({PRJ}!{c}9>0, {PRJ}!{c}88*({c}49/{PRJ}!{c}9), 0)"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 52: Closing balance
R = 52; sc(ws6, R, 2, "  = STRAIGHT-LINE Deferred Revenue Balance", font=bold_font, border=thin_border, fill=blue_light)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    if m == 1:
        f = f"={c}49-{c}50-{c}51"
    else:
        prev = get_column_letter(col - 1)
        f = f"={prev}{R}+{c}49-{c}50-{c}51"
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = blue_light
ws6.cell(row=R, column=2).fill = blue_light

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Recognized EBITDA under each method
# ────────────────────────────────────────────────────────────────────────────
R = 54; sc(ws6, R, 2, "EBITDA UNDER EACH RECOGNITION METHOD (incl. returns)", font=section_font)
note(ws6, R, 6, "Uses adjusted costs (COGS + returns + OpEx)")

# Row 55: Adjusted total costs (adjusted COGS + return overhead + OpEx)
R = 55; sc(ws6, R, 2, "Adjusted Total Costs (COGS + returns + OpEx)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col)
    cell.value = f"={PRJ}!{c}105+{PRJ}!{c}106+{PRJ}!{c}70"
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 56: Cash EBITDA (adjusted — net of refunds)
R = 56; sc(ws6, R, 2, "Cash Basis EBITDA (after returns)", font=bold_font, border=thin_border, fill=green_fill)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col)
    cell.value = f"={PRJ}!{c}104-{c}55"
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = green_fill
ws6.cell(row=R, column=2).fill = green_fill

# Row 57: Hybrid EBITDA
R = 57; sc(ws6, R, 2, "Hybrid EBITDA (after returns)", font=bold_font, border=thin_border, fill=orange_fill)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col)
    cell.value = f"={c}26-{PRJ}!{c}94-{c}55"
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = orange_fill
ws6.cell(row=R, column=2).fill = orange_fill

# Row 58: Straight-line EBITDA
R = 58; sc(ws6, R, 2, "Straight-Line EBITDA (after returns)", font=bold_font, border=thin_border, fill=blue_light)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col)
    cell.value = f"={c}31-{PRJ}!{c}94-{c}55"
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = blue_light
ws6.cell(row=R, column=2).fill = blue_light

# ────────────────────────────────────────────────────────────────────────────
# SECTION: Coverage Ratio — can the deferred pool cover future hosting?
# ────────────────────────────────────────────────────────────────────────────
R = 60; sc(ws6, R, 2, "ESCROW HEALTH — Can deferred revenue cover future hosting?", font=section_font)

# Row 61: Remaining hosting liability (net active charms after returns)
R = 61; sc(ws6, R, 2, "Est. Future Hosting Obligation (net active charms x monthly cost x lifetime)", font=normal_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    # Use net active charms from projections row 112 (adjusted for returns)
    f = (f"=({c}15*{PCC}!C56*$G$6 + {c}16*{PCC}!C56*$H$6 + {c}17*{PCC}!C56*$I$6)")
    cell = ws6.cell(row=R, column=col); cell.value = f
    cell.font = normal_font; cell.number_format = currency_whole; cell.border = thin_border

# Row 62: Hybrid deferred balance
R = 62; sc(ws6, R, 2, "Hybrid Deferred Revenue Balance (escrow pool)", font=bold_font, border=thin_border, fill=orange_fill)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col); cell.value = f"={c}46"
    cell.font = bold_font; cell.number_format = currency_whole; cell.border = thin_border; cell.fill = orange_fill
ws6.cell(row=R, column=2).fill = orange_fill

# Row 63: Coverage ratio (deferred balance / future obligation)
R = 63; sc(ws6, R, 2, "COVERAGE RATIO (deferred balance / hosting obligation)", font=bold_font, border=thin_border)
for m in range(1, 25):
    col = m + 2; c = get_column_letter(col)
    cell = ws6.cell(row=R, column=col)
    cell.value = f"=IF({c}61>0,{c}62/{c}61,0)"
    cell.font = bold_font; cell.number_format = pct_fmt; cell.border = thin_border
note(ws6, R, 6, "> 100% = escrow covers all future hosting. < 100% = shortfall risk.")

ws6.freeze_panes = "C5"


# ── Save ─────────────────────────────────────────────────────────────────────
output_path = r"c:\Users\appli\source\repos\MemoryCharm\MemoryCharm_Financial_Model.xlsx"
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print("Done!")
